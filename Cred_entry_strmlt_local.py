import streamlit as st
import pandas as pd
from datetime import datetime
import os
from pathlib import Path
import openpyxl
import json
from io import BytesIO

CASHIERS = sorted(["Misrak", "Emush", "Adanu", "Yemisrach", "Ejigayehu", "Tigist"])
BANKS = sorted([
    "Abay", "Amhara", "Awash", "Bank of Abyssinia", "Bunna",
    "CBE", "Dashen", "Enat", "Hibret", "Lion", "Nib", "Telebirr", "Wegagen", "Zemen"
])
HEADERS = ["ID", "Timestamp", "Cashier", "Bank", "Credit"]
SESSION_STATE_FILE = "session_state.json"

def get_today_str():
    
    return datetime.now().strftime("%Y-%m-%d")

def save_session_state():
  
    if st.session_state.get("current_aggregate_file") and st.session_state.get("first_entry_date"):
        state = {
            "aggregate_file": st.session_state.current_aggregate_file,
            "first_entry_date": st.session_state.first_entry_date
        }
        with open(SESSION_STATE_FILE, 'w') as f:
            json.dump(state, f)

def get_aggregate_excel_file():
  
    if Path(SESSION_STATE_FILE).exists():
        try:
            with open(SESSION_STATE_FILE, 'r') as f:
                state = json.load(f)
            st.session_state.current_aggregate_file = state.get("aggregate_file")
            st.session_state.first_entry_date = state.get("first_entry_date")
            return st.session_state.current_aggregate_file
        except (IOError, json.JSONDecodeError):
            if Path(SESSION_STATE_FILE).exists():
                os.remove(SESSION_STATE_FILE)

    Path(".").mkdir(exist_ok=True)
    base_filename = f"aggregate_{get_today_str()}"
    filename_to_use = f"{base_filename}.xlsx"
    counter = 1
    while Path(filename_to_use).exists():
        filename_to_use = f"{base_filename}_{counter}.xlsx"
        counter += 1

    st.session_state.current_aggregate_file = filename_to_use
    st.session_state.first_entry_date = get_today_str()
    save_session_state()
    return filename_to_use

def get_next_id(ws):
   
    if ws.max_row <= 1:
        return 1
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        cell_value = row[0]
        if cell_value is not None:
            try:
                current_id = int(cell_value)
                if current_id > max_id:
                    max_id = current_id
            except (ValueError, TypeError):
                continue
    return max_id + 1

def save_entry(entry):
   
    filepath = get_aggregate_excel_file()
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Entries"
        ws.append(HEADERS)

    entry_id = get_next_id(ws)
    entry["ID"] = entry_id
    
    data_row = [entry.get(h, "") for h in HEADERS]
    ws.append(data_row)
    wb.save(filepath)
    save_session_state()
    return entry

def remove_entry_from_excel(entry_id):
   
    filepath = get_aggregate_excel_file()
    if not os.path.exists(filepath):
        return False, "File not found."
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    row_to_delete = -1
    
    for row_index in range(2, ws.max_row + 1):
        cell_id = ws.cell(row=row_index, column=1).value
        if str(cell_id) == str(entry_id):
            row_to_delete = row_index
            break

    if row_to_delete != -1:
        ws.delete_rows(row_to_delete)
        wb.save(filepath)
        return True, f"Entry ID {entry_id} deleted successfully."
    
    wb.close()
    return False, f"Could not find entry ID {entry_id} in the file."

def load_data_from_excel():
  
    filepath = get_aggregate_excel_file()
    if not os.path.exists(filepath):
        return pd.DataFrame(columns=HEADERS)
    try:
        df = pd.read_excel(filepath, engine='openpyxl')
        if 'ID' in df.columns:
            df['ID'] = df['ID'].astype(str)
        return df
    except Exception as e:
        st.error(f"Failed to load data from {filepath}: {e}")
        return pd.DataFrame(columns=HEADERS)



def cashier_selection_page():
   
    st.header("Select Cashier to Continue")
    
    cols = st.columns(3)
    for i, cashier in enumerate(CASHIERS):
        with cols[i % 3]:
            if st.button(cashier, key=f"cashier_{cashier}", use_container_width=True):
                st.session_state.selected_cashier = cashier
                st.rerun()

def main_app_page():
   
    st.markdown("""
        <style>
            /* Reduce top padding */
            .main .block-container {
                padding-top: 2rem;
            }
            /* Style for bank buttons */
            .bank-buttons .stButton button {
                font-size: 1.2rem; /* Larger font size */
                font-weight: bold;   /* Bold font */
                height: 50px;
            }
        </style>
    """, unsafe_allow_html=True)

    cashier = st.session_state.selected_cashier

    
    top_col1, top_col2 = st.columns([1, 4])
    with top_col1:
        if st.button("⬅️ Change Cashier"):
            del st.session_state.selected_cashier
            if 'selected_bank' in st.session_state:
                del st.session_state.selected_bank
            st.rerun()
    with top_col2:
        st.markdown(f"## Welcome, **{cashier}**!")

    
    if 'selected_bank' not in st.session_state:
        st.session_state.selected_bank = None

   
    col1, col2 = st.columns([1.5, 2])

    
    with col1:
        st.subheader("1. Select a Bank")
        
        
        st.markdown('<div class="bank-buttons">', unsafe_allow_html=True)
        bank_cols = st.columns(2) 
        for i, bank in enumerate(BANKS):
            col = bank_cols[i % 2]
            is_selected = (bank == st.session_state.selected_bank)
            button_type = "primary" if is_selected else "secondary"
            if col.button(bank, key=f"bank_{bank}", use_container_width=True, type=button_type):
                st.session_state.selected_bank = bank
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

        st.subheader("2. Enter Amount & Submit")
        with st.form("entry_form", clear_on_submit=True):
            credit = st.number_input("Enter Credit Amount", min_value=0.01, format="%.2f", value=None)
            
            submitted = st.form_submit_button("Submit Entry", use_container_width=True)
            if submitted:
                bank = st.session_state.get("selected_bank")
                if not bank:
                    st.warning("Please select a bank.")
                elif credit is None:
                    st.warning("Please enter a credit amount.")
                else:
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    entry_data = { "Timestamp": timestamp, "Bank": bank, "Credit": credit, "Cashier": cashier }
                    try:
                        saved_entry = save_entry(entry_data)
                        st.success(f"Saved: ID {saved_entry['ID']} | {bank} - {credit:,.2f}")
                    except Exception as e:
                        st.error(f"Failed to save entry: {e}")

   
    with col2:
        st.subheader("Entries Log")
        
        df = load_data_from_excel()

        search_query = st.text_input("Search Entries", placeholder="Search by any column...")
        if search_query:
            mask = df.apply(lambda row: search_query.lower() in ' '.join(row.astype(str)).lower(), axis=1)
            filtered_df = df[mask]
        else:
            filtered_df = df
        
        display_df = filtered_df.sort_values(by="ID", ascending=False).reset_index(drop=True)

        st.dataframe(
            display_df, 
            use_container_width=True, 
            hide_index=True,
            key="entries_df",
            on_select="rerun",
            selection_mode="single-row"
        )

        st.markdown("---")
        st.subheader("Delete an Entry")
        st.write("Click a row in the table above to select it.")
        
        try:
            selection = st.session_state.entries_df["selection"]["rows"]
            if selection:
                selected_row_index = selection[0]
                selected_id = display_df.iloc[selected_row_index]["ID"]
                
                st.warning(f"You have selected Entry ID **{selected_id}** for deletion.")
                if st.button(f"Confirm Deletion of ID {selected_id}", type="primary", use_container_width=True):
                    success, message = remove_entry_from_excel(selected_id)
                    if success:
                        st.success(message)
                        st.rerun()
                    else:
                        st.error(message)
            else:
                st.info("No entry selected.")
        except (KeyError, IndexError):
             st.info("No entry selected.")

    
    st.markdown("---") 
    _, end_session_col = st.columns([4, 1]) 
    with end_session_col:
        if st.button("❌ End Session"):
            if Path(SESSION_STATE_FILE).exists():
                os.remove(SESSION_STATE_FILE)
            st.session_state.clear()
            st.success("Session ended. Please refresh the page to start a new one.")
            st.stop()



if __name__ == "__main__":
    st.set_page_config(page_title="Credit Entry", layout="wide")

    if 'selected_cashier' not in st.session_state:
        cashier_selection_page()
    else:
        main_app_page()

